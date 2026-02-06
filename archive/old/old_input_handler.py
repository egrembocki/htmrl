import csv
import json
import os
from typing import Any, Union

import numpy as np
from openpyxl import load_workbook

"""InputHandler singleton to pass a Data object to the Encoder layer. Implemented as a singleton layer handler for now, with methods to convert raw data to DataFrame, sequence, etc."""


class InputHandler:
    """
    Singleton InputHandler class to handle input data.

    """

    _instance = None
    """The single instance of the InputHandler class."""

    _data: Union[list[dict[str, Any]], list, np.ndarray, dict, str] = []
    """The input data of any type."""

    _hyperparameters: dict = {}
    """The hyperparameters associated with the input data."""

    # Getters, maybe use properties later
    def get_data(self) -> Union[list[dict[str, Any]], list, np.ndarray, dict, str]:
        """Getter for the data attribute"""
        return self._data

    def get_hyperparameters(self) -> dict:
        """Getter for the hyperparameters attribute"""
        return self._hyperparameters

    # main methods to handle input data processing

    def load_data(self, filepath: str) -> Union[list[dict[str, Any]], list, dict, str]:
        """Load data from a file based on file extension."""

        if not os.path.exists(filepath):
            raise FileNotFoundError(f"The file {filepath} does not exist.")

        file_extension = os.path.splitext(filepath)[1].lower()

        if file_extension == ".csv":
            print("Loading csv file:", file_extension, filepath)
            with open(filepath, "r", encoding="utf-8") as file:
                self._data = list(csv.DictReader(file))
            return self._data
        if file_extension in [".xls", ".xlsx"]:
            print("Loading excel file:", file_extension, filepath)
            workbook = load_workbook(filepath, read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                self._data = []
            else:
                headers = [
                    str(h) if h is not None else f"col_{idx}"
                    for idx, h in enumerate(rows[0])
                ]
                records = []
                for row in rows[1:]:
                    record = {headers[idx]: value for idx, value in enumerate(row)}
                    records.append(record)
                self._data = records
            return self._data
        if file_extension == ".json":
            print("Loading json file:", file_extension, filepath)
            with open(filepath, "r", encoding="utf-8") as file:
                self._data = json.load(file)
            return self._data
        if file_extension == ".txt":
            # setup context manager to read text file
            with open(filepath, "r") as file:
                self._data = file.readlines()
            return self._data

        raise ValueError(f"Unsupported file type: {file_extension}")

    def to_records(self, data: Union[list[dict[str, Any]], list, dict]) -> list[dict[str, Any]]:
        """Explicitly convert input data to a list of record dictionaries."""

        if isinstance(data, list):
            if not data:
                return []
            if isinstance(data[0], dict):
                return data
            return [{"value": entry} for entry in data]
        if isinstance(data, dict):
            return [data]
        raise TypeError("Unsupported data type for conversion to record list.")

    def raw_to_sequence(self) -> list:
        """Convert raw data to sequence (list)"""

        # Placeholder implementation; actual conversion logic will depend on data type
        # could be list, numpy array, dict, etc.

        if isinstance(self._data, list):
            return self._data
        else:
            # simple conversion to list, more logic is needed based on what the data is
            return [self._data]

    # validation methods

    def validate_data(self) -> bool:
        """Validate the input data"""

        # Placeholder implementation; actual validation logic will depend on data type and requirements

        is_valid = isinstance(self._data, (list, np.ndarray, dict, str))

        return is_valid

    def fill_missing_values(self, data: list[dict[str, Any]]) -> None:
        """Fill missing values in the input data"""

        # Placeholder implementation; actual logic will depend on data type and requirements

        if not data:
            return

        # Add more cases as needed for different data types

    def __new__(cls, *args, **kwargs):
        """Constructor -- Singleton pattern implementation."""

        if cls._instance is None:
            cls._instance = super(InputHandler, cls).__new__(cls)

        return cls._instance
