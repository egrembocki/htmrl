"""input handler stub - TDD"""

import csv
import json
import logging
from typing import Any, Union

import numpy as np
from openpyxl import load_workbook
from typing_extensions import Self


class InputHandler:
    """Stub InputHandler class for TDD purposes."""

    # Set up logging for this module
    logging.basicConfig(level=logging.INFO)

    _instance = None
    """The single instance of the InputHandler class."""

    _data: list[dict[str, Any]] = []
    """The input data as record dictionaries."""

    _raw_data: Union[str, bytearray, list, dict, np.ndarray] = ""
    """The raw input data of supported type."""

    def __new__(cls) -> Self:
        if cls._instance is None:
            cls._instance = super(InputHandler, cls).__new__(cls)
        return cls._instance

    def __init__(self):
        assert isinstance(self, InputHandler)
        self._data = []
        self._raw_data = ""
        logging.info("InputHandler initialized.")

    def load_data(self, filepath: str) -> Union[list[dict[str, Any]], list]:
        """Load data from a file into record dictionaries based on file extension."""
        assert isinstance(filepath, str), "File path must be a string"
        if filepath.endswith(".csv"):
            with open(filepath, "r", encoding="utf-8") as file:
                return list(csv.DictReader(file))
        if filepath.endswith(".xls") or filepath.endswith(".xlsx"):
            workbook = load_workbook(filepath, read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(h) if h is not None else f"col_{idx}" for idx, h in enumerate(rows[0])]
            records = []
            for row in rows[1:]:
                record = {headers[idx]: value for idx, value in enumerate(row)}
                records.append(record)
            return records
        if filepath.endswith(".json"):
            with open(filepath, "r", encoding="utf-8") as file:
                return json.load(file)
        if filepath.endswith(".txt"):
            with open(filepath, "r", encoding="utf-8") as file:
                return file.readlines()
        raise ValueError("Unsupported file type")

    def get_data(self) -> list[dict[str, Any]]:
        """Getter for the data attribute"""
        return self._data
