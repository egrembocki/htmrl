"""InputHandler for HTM pipeline. Normalize varied payloads into encoder-ready records."""

from __future__ import annotations

import csv
import datetime
import json
import os
from collections.abc import Mapping, Sequence
from typing import Any, ClassVar

import numpy as np
from openpyxl import load_workbook

from psu_capstone.input_layer.input_interface import InputInterface
from psu_capstone.log import logger


class InputHandler:
    """Build an input data manager."""

    __instance: ClassVar[InputHandler | None] = None
    """Singleton instance"""

    __interface: ClassVar[Any]  # may get removed in future updates -- sonarqube(python:S4487)
    """Interface for future extension"""

    _TEXT_EXTENSION: ClassVar[str] = ".txt"
    """File extension for text files."""

    def __new__(cls) -> "InputHandler":
        """Constructor ::  Singleton pattern implementation."""

        if cls.__instance is None:
            cls.__instance = super(InputHandler, cls).__new__(cls)

        return cls.__instance

    def __init__(self, data: Any = None) -> None:
        """Constructor :: called after new().

        Args:
            data (Any, optional): Initial data to load into the handler. Defaults to None.

        """

        self._data: list[dict[str, Any]]
        """The processed input data -> encoder-ready records."""

        self._columns: list[str] = []
        """Ordered column names for the current dataset."""

        self._appended_required_columns: set[str] = set()
        """Set of required columns."""

        self._required_columns_context: list[str] = []
        """Most-recent required_columns list passed to input_data (used for timestamp behavior)"""

    @classmethod
    def get_instance(cls) -> "InputHandler":
        """Static access method to get the singleton instance."""

        if cls.__instance is None:
            cls.__instance = InputHandler()
        return cls.__instance

    @property
    def data(self) -> list[dict[str, Any]]:
        return self._data

    @data.setter
    def data(self, data: list[dict[str, Any]]) -> None:
        self._data = data

    @property
    def interface(self) -> Any:
        return type(self).__interface

    @interface.setter
    def interface(self, interface: Any) -> None:
        type(self).__interface = interface

    def input_data(
        self, input_source: Any, required_columns: list[str] | None = None
    ) -> list[dict[str, Any]]:
        """Public :: exposed method to the user. Inputing data into the handler starts here

        raises:
            TypeError: If input_source is a path-like object.
            FileNotFoundError: If a file path is provided but the file does not exist.
            ValueError: If data validation fails.
        """

        self._appended_required_columns.clear()
        self._required_columns_context = required_columns or []

        if isinstance(input_source, bytes):
            input_source = bytearray(input_source)

        # Check if input is a file path
        if isinstance(input_source, os.PathLike):
            raise TypeError("Path-like objects must be converted to strings before ingestion.")

        if isinstance(input_source, str):
            file_extension = os.path.splitext(input_source)[1].lower()
            if os.path.exists(input_source):
                # Load data from file :: raw_data -> normalized_records
                raw_data = self._load_from_file(input_source)
                normalized_records = self._raw_to_sequence(raw_data)

                self._data = normalized_records

                if required_columns:
                    self._apply_required_columns(required_columns)
                elif self._data:
                    self._columns = list(self._data[0].keys())
                self._validate_data(required_columns)
                return self._data

            if file_extension in {
                ".csv",
                ".xlsx",
                ".xls",
                ".json",
                ".parquet",
                self._TEXT_EXTENSION,
            }:
                raise FileNotFoundError(f"No file found at {input_source}")

        normalized_records = self._raw_to_sequence(input_source)
        self._data = normalized_records
        if required_columns:
            self._apply_required_columns(required_columns)
        elif self._data:
            self._columns = list(self._data[0].keys())
        self._validate_data(required_columns)
        return self._data

    # return a np.ndarray from record lists
    def to_numpy(self, data: list[dict[str, Any]]) -> np.ndarray:
        """Convert record data to a numpy ndarray.

        Args:
            data (list[dict[str, Any]]): Input record list.
        Returns:
            np.ndarray: The converted numpy ndarray.
        """

        self._data = data
        validate_data = self._validate_data()  # type: ignore
        if not validate_data:
            raise ValueError("Data validation failed; cannot convert to numpy ndarray.")

        columns = self._columns or list(data[0].keys())
        matrix = [[row.get(col) for col in columns] for row in data]
        return np.asarray(matrix, dtype=np.float64)

    def _load_from_file(self, filepath: str) -> Any:
        """Read supported files into record-friendly structures."""

        try:
            logger.info("Loading data from %s", filepath)

            file_extension = os.path.splitext(filepath)[1].lower()

            if file_extension == ".csv":
                with open(filepath, "r", encoding="utf-8") as file:
                    return list(csv.DictReader(file))
            if file_extension == ".json":
                with open(filepath, "r", encoding="utf-8") as file:
                    return json.load(file)
            if file_extension == ".xlsx":
                workbook = load_workbook(filepath, read_only=True, data_only=True)
                sheet = workbook.active
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    return []
                headers = [
                    str(h) if h is not None else f"col_{idx}" for idx, h in enumerate(rows[0])
                ]
                records = []
                for row in rows[1:]:
                    record = {headers[idx]: value for idx, value in enumerate(row)}
                    records.append(record)
                return records
            if file_extension == ".xls":
                raise ValueError("Unsupported file type: .xls (requires optional xls reader).")
            if file_extension == self._TEXT_EXTENSION:
                with open(filepath, "r", encoding="utf-8") as file:
                    return [line for line in file.readlines()]
            raise ValueError(f"Unsupported file type: {file_extension}")

        except Exception as e:
            logger.error("Error loading file %s: %s", filepath, e)
            raise

    def _raw_to_sequence(self, data: Any) -> list[dict[str, Any]]:
        """Turn raw payloads into sequence-friendly records while flagging temporal values."""

        logger.info("converting to sequence")

        records, is_nested = self._coerce_records_for_sequence(data)
        self._fill_missing_values(records)
        normalized_records, contains_sequential = self._normalize_record_entries(records)

        if is_nested:
            logger.info("Detected multi-column input from nested iterables.")

        # If we did not detect temporal data, only inject a timestamp column when the caller
        # explicitly requested one (required_columns includes "timestamp").
        if not contains_sequential:
            if "timestamp" in self._required_columns_context:
                normalized_records = self._prepend_timestamp_column(normalized_records)
                logger.info("No temporal data detected; prepended timestamp column.")
            else:
                logger.info("No temporal data detected; continuing without timestamp column")

        return normalized_records

    def _coerce_records_for_sequence(self, data: Any) -> tuple[list[dict[str, Any]], bool]:
        """
        Coerce raw payloads into record lists while tracking whether the data was inherently
        multi-dimensional.

        Returns:
            tuple[list[dict[str, Any]], bool]: The coerced records and a flag indicating whether the
            input provided multiple columns (either originally or due to nested iterables).
        """

        if isinstance(data, list) and data and all(isinstance(item, Mapping) for item in data):
            return self._normalize_record_keys(data), True

        if isinstance(data, Mapping):
            return self._records_from_mapping(data), True

        iterable, is_nested = self._coerce_iterable_for_sequence(data)
        if is_nested:
            records = self._records_from_nested_iterable(iterable)
            return records, True

        return [{"value": item} for item in iterable], False

    def _coerce_iterable_for_sequence(self, data: Any) -> tuple[Sequence[Any], bool]:
        """Standardize iterables into lists and detect nested structures."""

        if isinstance(data, np.ndarray):
            iterable = data.tolist()
        elif isinstance(data, bytearray):
            iterable = data
        elif isinstance(data, bytes):
            iterable = bytearray(data)
        elif isinstance(data, list):
            iterable = data[:]
        elif isinstance(data, Sequence) and not isinstance(data, (str, bytes, bytearray)):
            iterable = list(data)
        elif isinstance(data, str):
            iterable = [data]
        elif isinstance(data, dict):
            iterable = list(data.values())
        else:
            raise TypeError(
                "Unsupported data type for conversion to sequence. "
                "Supported types: list, tuple, bytearray, bytes, numpy ndarray, or string."
            )

        is_nested = any(
            isinstance(item, Sequence) and not isinstance(item, (str, bytes, bytearray))
            for item in iterable
        )
        return iterable, is_nested

    def _prepend_timestamp_column(self, records: list[dict[str, Any]]) -> list[dict[str, Any]]:
        """Insert a leading timestamp field using the current time for each row."""

        base_time = datetime.datetime(1970, 1, 1)
        timestamp_values = [
            (base_time + datetime.timedelta(seconds=idx)).isoformat()
            for idx in range(len(records))
        ]
        updated_records = []
        for idx, record in enumerate(records):
            updated_record = {"timestamp": timestamp_values[idx], **record}
            updated_records.append(updated_record)
        return updated_records

    def _normalize_datetime_entry(self, value: object) -> tuple[object, bool]:
        """Return ISO-8601 strings for date-like values and note if a conversion occurred.

        Args:
            value: Scalar under inspection.

        Returns:
            tuple[object, bool]: Possibly-transformed value and whether a datetime was detected.
        """
        if isinstance(value, datetime.datetime):
            return value.isoformat(), True
        elif isinstance(value, datetime.date):
            # Avoid double conversion if already datetime
            return datetime.datetime.combine(value, datetime.time()).isoformat(), True
        elif isinstance(value, str):
            try:
                parsed = datetime.datetime.fromisoformat(value)
                return parsed.isoformat(), True
            except ValueError:
                return value, False
        else:
            return value, False

    def _fill_missing_values(self, data: Any) -> None:
        """Apply lightweight mean imputation for numeric containers when practical."""

        logger.info("filling missing values...")

        if isinstance(data, list):
            if not data:
                return
            if all(isinstance(item, Mapping) for item in data):
                numeric_columns: dict[str, list[float]] = {}
                for row in data:
                    for key, value in row.items():
                        if isinstance(value, (int, float, np.number)) and value is not None:
                            numeric_columns.setdefault(key, []).append(float(value))
                column_means = {
                    key: (sum(values) / len(values)) if values else None
                    for key, values in numeric_columns.items()
                }
                for row in data:
                    for key, mean_value in column_means.items():
                        if row.get(key) is None and mean_value is not None:
                            row[key] = mean_value
            else:
                numeric_values = [v for v in data if isinstance(v, (int, float, np.number))]
                mean_value = (sum(numeric_values) / len(numeric_values)) if numeric_values else None
                if mean_value is not None:
                    for idx, value in enumerate(data):
                        if value is None:
                            data[idx] = mean_value
        elif isinstance(data, dict):
            for key, value in data.items():
                if isinstance(value, list):
                    numeric_values = [v for v in value if isinstance(v, (int, float, np.number))]
                    mean_value = (
                        sum(numeric_values) / len(numeric_values) if numeric_values else None
                    )
                    data[key] = [mean_value if v is None else v for v in value]
        elif isinstance(data, np.ndarray):
            if np.issubdtype(data.dtype, np.number):
                mean_value = np.nanmean(data)
                inds = np.nonzero(np.isnan(data))
                data[inds] = mean_value
        elif isinstance(data, tuple):
            logger.info("tuples are immutable; no missing value handling applied")
        else:
            logger.info("no missing value handling for type: %s", type(data))

    def _validate_data(self, required_columns: list[str] | None = None) -> bool:
        """Verify structure, NaN patterns, duplicates, and caller-specified schemas."""

        logger.info("validating data...")
        # Check type
        if not isinstance(self._data, list) or (
            self._data and not all(isinstance(row, Mapping) for row in self._data)
        ):
            raise ValueError("data is not a list of records.")

        # Check empty
        if not self._data:
            raise ValueError("Record list is empty")

        columns = self._columns or list(self._data[0].keys())

        # Check for duplicate columns
        if len(set(columns)) != len(columns):
            raise ValueError("Record list has duplicate column names.")

        # Check for all-None columns
        nan_cols = [
            col
            for col in columns
            if all(row.get(col) is None for row in self._data)
            and col not in self._appended_required_columns
        ]
        if nan_cols:
            raise ValueError(f"Columns with all None values: {nan_cols}")

        if self._appended_required_columns:
            logger.info(
                "Appended required columns with placeholder values: %s",
                sorted(self._appended_required_columns),
            )

        # Check for required columns
        if required_columns:
            missing = [col for col in required_columns if col not in columns]
            if missing:
                raise ValueError(f"Missing required columns: {missing}")

        logger.info("data has been validated")
        return True

    def _apply_required_columns(self, required_columns: list[str] = []) -> None:
        """Align columns to the requested order and append NA placeholders for missing names."""

        if not required_columns:
            return

        # Required column lists must be unique; duplicates would create invalid record schemas.
        if len(set(required_columns)) != len(required_columns):
            seen: set[str] = set()
            duplicates = [c for c in required_columns if (c in seen) or seen.add(c)]
            raise ValueError(f"Duplicate entries in required_columns: {sorted(set(duplicates))}")

        self._columns = required_columns[:]
        normalized_records = []
        current_columns = list(self._data[0].keys()) if self._data else []
        rename_count = min(len(required_columns), len(current_columns))
        for record in self._data:
            normalized_record: dict[str, Any] = {}
            for idx, column in enumerate(required_columns):
                if idx < rename_count:
                    normalized_record[column] = record.get(current_columns[idx])
                else:
                    normalized_record[column] = None
                    self._appended_required_columns.add(column)
            normalized_records.append(normalized_record)
        self._data = normalized_records

    def _validate_sequence(self, data: Any) -> bool:
        """Placeholder sequence check that only ensures the payload is iterable."""

        # New Algorithm to peek at data set and find a periodic sequence in it

        logger.info("validating sequence...")

        # Treat common containers as acceptable "sequence-like" inputs for our pipeline.
        # Record lists and numpy arrays are valid sequences of rows.
        if isinstance(data, (list, tuple, np.ndarray)):
            return True

        logger.warning("Input data is not a sequence.")
        return False

    def _records_from_mapping(self, data: Mapping[str, Any]) -> list[dict[str, Any]]:
        """Create records from mapping inputs (dicts of lists or scalar dicts)."""

        values = list(data.values())
        if values and all(isinstance(value, list) for value in values):
            max_len = max(len(value) for value in values)
            records = []
            for idx in range(max_len):
                row = {
                    key: (col_values[idx] if idx < len(col_values) else None)
                    for key, col_values in data.items()
                }
                records.append(row)
            return records
        return [dict(data)]

    def _records_from_nested_iterable(self, iterable: list[Any]) -> list[dict[str, Any]]:
        """Create records from nested iterables like list[list[Any]]."""

        num_cols = max(len(row) for row in iterable) if iterable else 0
        columns = [f"col_{idx}" for idx in range(num_cols)]
        records = []
        for row in iterable:
            row_values = list(row)
            record = {
                columns[idx]: (row_values[idx] if idx < len(row_values) else None)
                for idx in range(len(columns))
            }
            records.append(record)
        return records

    def _normalize_record_keys(self, records: list[Mapping[str, Any]]) -> list[dict[str, Any]]:
        """Ensure all record dicts share the same keys."""

        ordered_keys: list[str] = []
        seen: set[str] = set()
        for record in records:
            for key in record.keys():
                if key not in seen:
                    ordered_keys.append(key)
                    seen.add(key)
        normalized = []
        for record in records:
            normalized.append({key: record.get(key) for key in ordered_keys})
        return normalized

    def _normalize_record_entries(
        self, records: list[dict[str, Any]]
    ) -> tuple[list[dict[str, Any]], bool]:
        """Normalize each value and report whether datetime-like content was encountered."""

        self._validate_sequence(records)
        contains_sequential = False

        def _normalize(value: Any) -> Any:
            nonlocal contains_sequential
            normalized, is_date = self._normalize_datetime_entry(value)
            contains_sequential = contains_sequential or is_date
            return normalized

        normalized_records = [
            {key: _normalize(value) for key, value in record.items()} for record in records
        ]
        return normalized_records, contains_sequential
